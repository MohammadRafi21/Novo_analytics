import json, os, csv as csv_mod
from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.conf import settings

try:
    import numpy as np
    import pandas as pd
    from sklearn.cluster import KMeans
    from sklearn.preprocessing import StandardScaler
    from sklearn.metrics import silhouette_score
    from sklearn.decomposition import PCA
    ML_AVAILABLE = True
except ImportError:
    ML_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ── Raw datasets for k-means ──────────────────────────────────────────────────
SUPPLIER_DATASET = [
    {"name":"Hindalco Industries","country":"India","capacity_mt":2500000,"price_usd_kg":2.10,"reliability_score":9.2,"lead_time_days":25,"compliance_score":9.0,"export_growth_pct":18,"anti_dumping_risk":0,"freight_usd_kg":0.22},
    {"name":"Vedanta Aluminium","country":"India","capacity_mt":1900000,"price_usd_kg":1.95,"reliability_score":8.5,"lead_time_days":28,"compliance_score":8.5,"export_growth_pct":15,"anti_dumping_risk":0,"freight_usd_kg":0.22},
    {"name":"NALCO","country":"India","capacity_mt":460000,"price_usd_kg":1.85,"reliability_score":8.8,"lead_time_days":30,"compliance_score":8.8,"export_growth_pct":12,"anti_dumping_risk":0,"freight_usd_kg":0.22},
    {"name":"BALCO","country":"India","capacity_mt":245000,"price_usd_kg":2.00,"reliability_score":8.0,"lead_time_days":32,"compliance_score":8.2,"export_growth_pct":10,"anti_dumping_risk":0,"freight_usd_kg":0.23},
    {"name":"Century Aluminium","country":"India","capacity_mt":120000,"price_usd_kg":2.05,"reliability_score":7.8,"lead_time_days":35,"compliance_score":7.8,"export_growth_pct":9,"anti_dumping_risk":0,"freight_usd_kg":0.23},
    {"name":"Rajhans Metals","country":"India","capacity_mt":45000,"price_usd_kg":1.90,"reliability_score":7.2,"lead_time_days":40,"compliance_score":7.0,"export_growth_pct":8,"anti_dumping_risk":0,"freight_usd_kg":0.24},
    {"name":"GreenTech Aluminium","country":"India","capacity_mt":12000,"price_usd_kg":1.75,"reliability_score":7.5,"lead_time_days":38,"compliance_score":7.5,"export_growth_pct":11,"anti_dumping_risk":0,"freight_usd_kg":0.24},
    {"name":"Aluplex India","country":"India","capacity_mt":28000,"price_usd_kg":1.88,"reliability_score":6.9,"lead_time_days":42,"compliance_score":6.9,"export_growth_pct":7,"anti_dumping_risk":0,"freight_usd_kg":0.24},
    {"name":"Aluminium Corp of China","country":"China","capacity_mt":3800000,"price_usd_kg":2.30,"reliability_score":9.0,"lead_time_days":16,"compliance_score":6.0,"export_growth_pct":-5,"anti_dumping_risk":8,"freight_usd_kg":0.32},
    {"name":"Zhongwang Group","country":"China","capacity_mt":2100000,"price_usd_kg":2.25,"reliability_score":8.5,"lead_time_days":17,"compliance_score":5.5,"export_growth_pct":-8,"anti_dumping_risk":9,"freight_usd_kg":0.32},
    {"name":"Nanshan Aluminium","country":"China","capacity_mt":1600000,"price_usd_kg":2.20,"reliability_score":8.2,"lead_time_days":18,"compliance_score":5.8,"export_growth_pct":-6,"anti_dumping_risk":8,"freight_usd_kg":0.33},
    {"name":"Alnan Aluminium","country":"China","capacity_mt":900000,"price_usd_kg":2.15,"reliability_score":7.8,"lead_time_days":18,"compliance_score":5.5,"export_growth_pct":-7,"anti_dumping_risk":8,"freight_usd_kg":0.33},
    {"name":"Henan Zhongfu","country":"China","capacity_mt":650000,"price_usd_kg":2.10,"reliability_score":7.5,"lead_time_days":19,"compliance_score":5.2,"export_growth_pct":-9,"anti_dumping_risk":9,"freight_usd_kg":0.33},
    {"name":"Thai Summit","country":"Thailand","capacity_mt":180000,"price_usd_kg":2.18,"reliability_score":7.8,"lead_time_days":20,"compliance_score":7.8,"export_growth_pct":5,"anti_dumping_risk":2,"freight_usd_kg":0.28},
    {"name":"TIS Group","country":"Thailand","capacity_mt":95000,"price_usd_kg":2.22,"reliability_score":7.5,"lead_time_days":22,"compliance_score":7.5,"export_growth_pct":4,"anti_dumping_risk":2,"freight_usd_kg":0.28},
    {"name":"Poonsab Aluminium","country":"Thailand","capacity_mt":60000,"price_usd_kg":2.15,"reliability_score":7.2,"lead_time_days":23,"compliance_score":7.2,"export_growth_pct":3,"anti_dumping_risk":2,"freight_usd_kg":0.29},
    {"name":"POSCO Aluminium","country":"South Korea","capacity_mt":420000,"price_usd_kg":2.45,"reliability_score":9.2,"lead_time_days":14,"compliance_score":8.8,"export_growth_pct":2,"anti_dumping_risk":1,"freight_usd_kg":0.30},
    {"name":"Dongyang Tinplate","country":"South Korea","capacity_mt":210000,"price_usd_kg":2.50,"reliability_score":8.8,"lead_time_days":15,"compliance_score":8.5,"export_growth_pct":1,"anti_dumping_risk":1,"freight_usd_kg":0.30},
    {"name":"Malaysia Smelting","country":"Malaysia","capacity_mt":145000,"price_usd_kg":2.12,"reliability_score":7.6,"lead_time_days":20,"compliance_score":7.6,"export_growth_pct":6,"anti_dumping_risk":2,"freight_usd_kg":0.27},
    {"name":"Press Metal Aluminium","country":"Malaysia","capacity_mt":220000,"price_usd_kg":2.08,"reliability_score":8.0,"lead_time_days":19,"compliance_score":8.0,"export_growth_pct":8,"anti_dumping_risk":1,"freight_usd_kg":0.27},
]

CONSTRUCTION_SEGMENTS = [
    {"segment":"Residential High-Rise","annual_demand_mt":85000,"market_share_pct":24.3,"growth_rate_pct":6.2,"avg_project_size_aud":45000000,"price_sensitivity":8,"compliance_priority":6,"speed_priority":7,"volume_per_order_mt":450},
    {"segment":"Commercial Office","annual_demand_mt":62000,"market_share_pct":17.8,"growth_rate_pct":3.5,"avg_project_size_aud":120000000,"price_sensitivity":6,"compliance_priority":9,"speed_priority":5,"volume_per_order_mt":620},
    {"segment":"Infrastructure","annual_demand_mt":58000,"market_share_pct":16.6,"growth_rate_pct":8.1,"avg_project_size_aud":500000000,"price_sensitivity":5,"compliance_priority":10,"speed_priority":4,"volume_per_order_mt":1200},
    {"segment":"Industrial Facilities","annual_demand_mt":44000,"market_share_pct":12.6,"growth_rate_pct":4.8,"avg_project_size_aud":80000000,"price_sensitivity":7,"compliance_priority":8,"speed_priority":6,"volume_per_order_mt":380},
    {"segment":"Residential Mid-Rise","annual_demand_mt":38000,"market_share_pct":10.9,"growth_rate_pct":9.3,"avg_project_size_aud":12000000,"price_sensitivity":9,"compliance_priority":5,"speed_priority":8,"volume_per_order_mt":180},
    {"segment":"Retail & Hospitality","annual_demand_mt":27000,"market_share_pct":7.7,"growth_rate_pct":2.1,"avg_project_size_aud":8000000,"price_sensitivity":8,"compliance_priority":6,"speed_priority":9,"volume_per_order_mt":90},
    {"segment":"Education & Health","annual_demand_mt":22000,"market_share_pct":6.3,"growth_rate_pct":5.5,"avg_project_size_aud":35000000,"price_sensitivity":6,"compliance_priority":9,"speed_priority":5,"volume_per_order_mt":220},
    {"segment":"Renovation & Fitout","annual_demand_mt":13000,"market_share_pct":3.7,"growth_rate_pct":7.8,"avg_project_size_aud":500000,"price_sensitivity":9,"compliance_priority":4,"speed_priority":9,"volume_per_order_mt":25},
    {"segment":"Mining Infrastructure","annual_demand_mt":19000,"market_share_pct":5.4,"growth_rate_pct":11.2,"avg_project_size_aud":250000000,"price_sensitivity":4,"compliance_priority":8,"speed_priority":5,"volume_per_order_mt":850},
    {"segment":"Aged Care Facilities","annual_demand_mt":8000,"market_share_pct":2.3,"growth_rate_pct":12.5,"avg_project_size_aud":18000000,"price_sensitivity":7,"compliance_priority":9,"speed_priority":6,"volume_per_order_mt":95},
]

CLUSTER_LABELS = {
    0:{"name":"Premium reliable suppliers","color":"#0d6efd","description":"High reliability and compliance scores. Best for government, infrastructure, and compliance-critical projects."},
    1:{"name":"Cost-competitive India suppliers","color":"#198754","description":"Lowest price, zero anti-dumping risk, growing export capacity. Primary sourcing recommendation under AI-ECTA."},
    2:{"name":"High-risk China incumbents","color":"#dc3545","description":"Large capacity but active anti-dumping duties (ADN 2023/051) and declining trade policy scores. Avoid for new contracts."},
    3:{"name":"Regional niche suppliers","color":"#fd7e14","description":"Mid-tier capacity from Thailand, Malaysia. Balanced price-risk profile. Good diversification option."},
}

SEGMENT_CLUSTER_LABELS = {
    0:{"name":"High-value compliance buyers","color":"#0d6efd","description":"Large projects, low price sensitivity, high compliance need. Match with Cluster 0 (Premium reliable) suppliers."},
    1:{"name":"SME price-sensitive buyers","color":"#198754","description":"High price sensitivity, moderate volume. Best match for Cluster 1 (Cost-competitive India) suppliers."},
    2:{"name":"Mega-project infrastructure","color":"#6f42c1","description":"Enormous volume, long lead times acceptable, strict compliance. Hindalco/Vedanta ideal match."},
    3:{"name":"Fast-turnaround growth segments","color":"#fd7e14","description":"High growth, urgent delivery, smaller orders. Mixed India + regional supplier strategy recommended."},
}

OPPORTUNITY_DATA = {
    "India":{"flag":"IN","color":"#ff9933","years":{2019:{"price_competitiveness":7.2,"supply_reliability":6.0,"trade_policy":7.5,"logistics_ease":5.5,"market_growth":6.8,"ai_platform_fit":6.0},2020:{"price_competitiveness":7.5,"supply_reliability":6.2,"trade_policy":7.8,"logistics_ease":5.8,"market_growth":6.5,"ai_platform_fit":6.5},2021:{"price_competitiveness":7.8,"supply_reliability":6.5,"trade_policy":8.0,"logistics_ease":6.2,"market_growth":7.2,"ai_platform_fit":7.0},2022:{"price_competitiveness":8.2,"supply_reliability":7.0,"trade_policy":8.5,"logistics_ease":6.5,"market_growth":7.8,"ai_platform_fit":7.5},2023:{"price_competitiveness":8.8,"supply_reliability":7.5,"trade_policy":9.0,"logistics_ease":7.0,"market_growth":8.5,"ai_platform_fit":8.5},2024:{"price_competitiveness":9.2,"supply_reliability":8.0,"trade_policy":9.2,"logistics_ease":7.5,"market_growth":9.0,"ai_platform_fit":9.0}}},
    "China":{"flag":"CN","color":"#de2910","years":{2019:{"price_competitiveness":9.0,"supply_reliability":9.2,"trade_policy":6.0,"logistics_ease":8.5,"market_growth":5.0,"ai_platform_fit":5.0},2020:{"price_competitiveness":9.0,"supply_reliability":9.0,"trade_policy":5.5,"logistics_ease":8.2,"market_growth":4.8,"ai_platform_fit":4.5},2021:{"price_competitiveness":8.5,"supply_reliability":8.8,"trade_policy":5.0,"logistics_ease":8.0,"market_growth":4.5,"ai_platform_fit":4.0},2022:{"price_competitiveness":8.0,"supply_reliability":8.5,"trade_policy":4.0,"logistics_ease":7.5,"market_growth":4.0,"ai_platform_fit":3.5},2023:{"price_competitiveness":7.0,"supply_reliability":8.0,"trade_policy":2.5,"logistics_ease":7.0,"market_growth":3.5,"ai_platform_fit":3.0},2024:{"price_competitiveness":6.5,"supply_reliability":7.5,"trade_policy":2.0,"logistics_ease":6.8,"market_growth":3.0,"ai_platform_fit":2.5}}},
    "Thailand":{"flag":"TH","color":"#0033a0","years":{2019:{"price_competitiveness":7.0,"supply_reliability":7.0,"trade_policy":7.0,"logistics_ease":7.2,"market_growth":5.5,"ai_platform_fit":5.5},2020:{"price_competitiveness":7.0,"supply_reliability":6.8,"trade_policy":7.2,"logistics_ease":7.0,"market_growth":5.0,"ai_platform_fit":5.5},2021:{"price_competitiveness":7.2,"supply_reliability":7.0,"trade_policy":7.3,"logistics_ease":7.2,"market_growth":5.5,"ai_platform_fit":6.0},2022:{"price_competitiveness":7.0,"supply_reliability":7.0,"trade_policy":7.0,"logistics_ease":7.0,"market_growth":5.5,"ai_platform_fit":6.0},2023:{"price_competitiveness":7.2,"supply_reliability":7.2,"trade_policy":7.5,"logistics_ease":7.3,"market_growth":6.0,"ai_platform_fit":6.5},2024:{"price_competitiveness":7.5,"supply_reliability":7.5,"trade_policy":7.8,"logistics_ease":7.5,"market_growth":6.5,"ai_platform_fit":7.0}}},
    "South Korea":{"flag":"KR","color":"#003478","years":{2019:{"price_competitiveness":6.5,"supply_reliability":8.5,"trade_policy":7.5,"logistics_ease":8.0,"market_growth":5.0,"ai_platform_fit":6.5},2020:{"price_competitiveness":6.5,"supply_reliability":8.5,"trade_policy":7.5,"logistics_ease":8.0,"market_growth":5.0,"ai_platform_fit":6.5},2021:{"price_competitiveness":6.3,"supply_reliability":8.5,"trade_policy":7.5,"logistics_ease":8.0,"market_growth":5.2,"ai_platform_fit":6.5},2022:{"price_competitiveness":6.0,"supply_reliability":8.2,"trade_policy":7.5,"logistics_ease":7.8,"market_growth":5.0,"ai_platform_fit":6.5},2023:{"price_competitiveness":6.0,"supply_reliability":8.0,"trade_policy":7.5,"logistics_ease":7.8,"market_growth":5.0,"ai_platform_fit":6.5},2024:{"price_competitiveness":6.0,"supply_reliability":8.0,"trade_policy":7.5,"logistics_ease":7.8,"market_growth":5.0,"ai_platform_fit":6.5}}},
    "Malaysia":{"flag":"MY","color":"#cc0001","years":{2019:{"price_competitiveness":7.0,"supply_reliability":6.5,"trade_policy":7.5,"logistics_ease":7.5,"market_growth":5.5,"ai_platform_fit":5.5},2020:{"price_competitiveness":7.0,"supply_reliability":6.5,"trade_policy":7.5,"logistics_ease":7.3,"market_growth":5.0,"ai_platform_fit":5.5},2021:{"price_competitiveness":7.2,"supply_reliability":6.8,"trade_policy":7.8,"logistics_ease":7.5,"market_growth":5.5,"ai_platform_fit":6.0},2022:{"price_competitiveness":7.3,"supply_reliability":7.0,"trade_policy":7.8,"logistics_ease":7.5,"market_growth":5.8,"ai_platform_fit":6.2},2023:{"price_competitiveness":7.5,"supply_reliability":7.2,"trade_policy":8.0,"logistics_ease":7.5,"market_growth":6.0,"ai_platform_fit":6.5},2024:{"price_competitiveness":7.5,"supply_reliability":7.5,"trade_policy":8.0,"logistics_ease":7.8,"market_growth":6.2,"ai_platform_fit":6.8}}},
    "Vietnam":{"flag":"VN","color":"#da251d","years":{2019:{"price_competitiveness":7.5,"supply_reliability":5.5,"trade_policy":7.0,"logistics_ease":6.0,"market_growth":6.5,"ai_platform_fit":5.0},2020:{"price_competitiveness":7.8,"supply_reliability":5.8,"trade_policy":7.2,"logistics_ease":6.2,"market_growth":6.8,"ai_platform_fit":5.5},2021:{"price_competitiveness":8.0,"supply_reliability":6.0,"trade_policy":7.5,"logistics_ease":6.5,"market_growth":7.2,"ai_platform_fit":6.0},2022:{"price_competitiveness":8.0,"supply_reliability":6.2,"trade_policy":7.5,"logistics_ease":6.5,"market_growth":7.0,"ai_platform_fit":6.0},2023:{"price_competitiveness":8.2,"supply_reliability":6.5,"trade_policy":7.8,"logistics_ease":6.8,"market_growth":7.5,"ai_platform_fit":6.5},2024:{"price_competitiveness":8.3,"supply_reliability":6.8,"trade_policy":8.0,"logistics_ease":7.0,"market_growth":7.8,"ai_platform_fit":7.0}}},
}
DIM_WEIGHTS = {"price_competitiveness":0.28,"supply_reliability":0.20,"trade_policy":0.22,"logistics_ease":0.12,"market_growth":0.10,"ai_platform_fit":0.08}

def compute_opp_score(dims):
    return round(sum(dims[d]*DIM_WEIGHTS[d] for d in DIM_WEIGHTS), 2)

INDIA_PRICE_DATA = {"quarters":["Q1'22","Q2'22","Q3'22","Q4'22","Q1'23","Q2'23","Q3'23","Q4'23","Q1'24","Q2'24","Q3'24","Q4'24"],"china_fob":[2.55,2.60,2.65,2.70,2.72,2.78,2.80,2.85,2.88,2.92,2.95,3.00],"india_fob":[2.10,2.12,2.15,2.18,2.20,2.22,2.25,2.28,2.30,2.33,2.35,2.38],"freight_china":[0.28,0.28,0.30,0.30,0.31,0.31,0.32,0.32,0.33,0.33,0.34,0.34],"freight_india":[0.22,0.22,0.23,0.23,0.24,0.24,0.25,0.25,0.25,0.26,0.26,0.26],"aud_usd":[0.72,0.70,0.68,0.67,0.68,0.66,0.65,0.64,0.65,0.66,0.65,0.63],"antidumping_china":[0.00,0.00,0.00,0.00,0.00,0.00,0.15,0.18,0.20,0.20,0.22,0.25]}

DEFAULT_DATASETS = {
    "au_imports":{"label":"AU Aluminium Imports 2024","source":"Australian Aluminium Council / ABS","type":"real","headers":["Product","H1 Weight (t)","H2 Weight (t)","Total Weight (t)","Total Value (AUD)"],"rows":[["Sheet & Plate","100,289","100,883","201,172","A$1,015,931,763"],["Extrusions","46,211","45,268","91,480","A$544,860,652"],["Unwrought Alloys","25,167","24,193","49,360","A$190,948,799"],["Alumina","76,000 kt","102,000 kt","178,000 kt","A$130,597,375"],["Foil","7,757","8,691","16,448","A$123,217,244"],["Wire","1,552","1,538","3,090","A$19,134,043"],["Unwrought (unalloyed)","119","340","459","A$2,325,156"],["Scrap","759","738","1,497","A$3,021,176"]]},
    "lme_prices":{"label":"LME Aluminium Price History","source":"LME / World Bank / Statista","type":"real","headers":["Year","LME Avg (USD/t)","AUD/USD Rate","AUD/t (est)","YoY Change","Key Event"],"rows":[["2019","$1,794","0.695","$2,581","—","Trade war uncertainty"],["2020","$1,704","0.691","$2,466","-5.0%","COVID-19 demand crash"],["2021","$2,472","0.752","$3,288","+45.1%","Post-COVID surge"],["2022","$2,710","0.694","$3,905","+9.6%","Russia-Ukraine war"],["2023","$2,250","0.660","$3,409","-17.0%","Post-crisis correction"],["2024","$2,419","0.650","$3,722","+7.5%","Recovery phase"]]},
    "au_partners":{"label":"AU Trade Partners","source":"ABS International Trade Supplementary","type":"real","headers":["Country","FY18-19 (AUD B)","FY20-21 (AUD B)","FY22-23 (AUD B)","FY23-24 (AUD B)"],"rows":[["China","81.6","88.7","113.7","112.6"],["USA","27.8","28.5","49.7","42.4"],["Japan","22.8","19.2","32.4","34.3"],["South Korea","25.8","26.3","35.4","37.9"],["India (est)","8.2","9.1","16.3","18.5"],["Thailand (est)","10.1","9.8","14.2","15.6"]]},
    "supplier_kmeans":{"label":"Supplier Dataset (K-Means Ready)","source":"Industry Reports + NOVO Model","type":"ml_ready","headers":["Supplier","Country","Capacity (MT)","Price USD/kg","Reliability","Lead Time (d)","Compliance","Export Growth %","Anti-Dumping Risk","Freight USD/kg"],"rows":[[s["name"],s["country"],str(s["capacity_mt"]),str(s["price_usd_kg"]),str(s["reliability_score"]),str(s["lead_time_days"]),str(s["compliance_score"]),str(s["export_growth_pct"]),str(s["anti_dumping_risk"]),str(s["freight_usd_kg"])] for s in SUPPLIER_DATASET]},
    "construction_kmeans":{"label":"Construction Segments (K-Means Ready)","source":"Industry Analysis + NOVO Model","type":"ml_ready","headers":["Segment","Demand (MT)","Market Share %","Growth %","Price Sensitivity","Compliance Priority","Speed Priority","Volume/Order MT"],"rows":[[s["segment"],str(s["annual_demand_mt"]),str(s["market_share_pct"]),str(s["growth_rate_pct"]),str(s["price_sensitivity"]),str(s["compliance_priority"]),str(s["speed_priority"]),str(s["volume_per_order_mt"])] for s in CONSTRUCTION_SEGMENTS]},
    "price_comparison":{"label":"China vs India Price Comparison","source":"Market Analysis (Estimated)","type":"estimated","headers":["Quarter","China FOB","India FOB","China Total AUD/kg","India Total AUD/kg","India Saving %"],"rows":[["Q1 2022","2.55","2.10","3.87","3.20","17.3%"],["Q2 2022","2.60","2.12","3.95","3.23","18.2%"],["Q3 2022","2.65","2.15","4.02","3.27","18.7%"],["Q4 2022","2.70","2.18","4.10","3.31","19.3%"],["Q1 2023","2.72","2.20","4.13","3.34","19.1%"],["Q2 2023","2.78","2.22","4.22","3.37","20.1%"],["Q3 2023","2.80","2.25","4.25","3.42","19.5%"],["Q4 2023","2.85","2.28","4.33","3.46","20.1%"],["Q1 2024","2.88","2.30","4.37","3.49","20.1%"],["Q2 2024","2.92","2.33","4.43","3.54","20.1%"],["Q3 2024","2.95","2.35","4.48","3.57","20.3%"],["Q4 2024","3.00","2.38","4.55","3.61","20.7%"]]},
    "ai_cost_benefit":{"label":"AI Platform Cost-Benefit","source":"Project Brief + Analysis","type":"estimated","headers":["Cost Category","Without AI (AUD)","With AI (AUD)","Saving (AUD)","Saving %","Notes"],"rows":[["Procurement (per 100MT)","$420,000","$388,000","$32,000","7.6%","5-8% reduction per brief"],["Freight & Logistics","$35,000","$28,000","$7,000","20.0%","Route optimisation"],["Customs & Compliance","$18,000","$9,000","$9,000","50.0%","Automated docs"],["Supplier Vetting","$12,000","$3,000","$9,000","75.0%","AI risk scoring"],["Trade Finance","$22,000","$17,000","$5,000","22.7%","AI credit scoring"],["Project Delay","$45,000","$18,000","$27,000","60.0%","Predictive logistics"],["Currency Risk","$8,000","$4,000","$4,000","50.0%","Hedging AI"],["Manual Admin","$15,000","$4,000","$11,000","73.3%","Automation"],["TOTAL","$575,000","$471,000","$104,000","18.1%","Combined saving"]]},
}
UPLOADED_DATASETS = {}

# ── K-Means engine ────────────────────────────────────────────────────────────
def run_kmeans_supplier(k=4):
    if not ML_AVAILABLE:
        return {"error":"scikit-learn not installed"}
    features = ["price_usd_kg","reliability_score","lead_time_days","compliance_score","anti_dumping_risk","freight_usd_kg"]
    df = pd.DataFrame(SUPPLIER_DATASET)
    X = df[features].values
    scaler = StandardScaler()
    X_sc = scaler.fit_transform(X)
    km = KMeans(n_clusters=k, random_state=42, n_init=10)
    labels = km.fit_predict(X_sc)
    sil = float(silhouette_score(X_sc, labels))
    pca = PCA(n_components=2, random_state=42)
    coords = pca.fit_transform(X_sc)
    inertias = []
    for ki in range(2,9):
        km2 = KMeans(n_clusters=ki, random_state=42, n_init=10)
        km2.fit(X_sc)
        inertias.append(round(float(km2.inertia_),2))
    cluster_stats = []
    for c in range(k):
        mask = labels==c
        m = df[mask]
        li = CLUSTER_LABELS.get(c,{"name":f"Cluster {c}","color":"#888888","description":""})
        cluster_stats.append({"id":int(c),"name":li["name"],"color":li["color"],"description":li["description"],"count":int(mask.sum()),"suppliers":m["name"].tolist(),"countries":m["country"].unique().tolist(),"avg_price":round(float(m["price_usd_kg"].mean()),3),"avg_reliability":round(float(m["reliability_score"].mean()),2),"avg_compliance":round(float(m["compliance_score"].mean()),2),"avg_lead_time":round(float(m["lead_time_days"].mean()),1),"avg_anti_dumping":round(float(m["anti_dumping_risk"].mean()),2),"avg_freight":round(float(m["freight_usd_kg"].mean()),3)})
    points = []
    for i,(row,lbl,coord) in enumerate(zip(SUPPLIER_DATASET,labels,coords)):
        li = CLUSTER_LABELS.get(int(lbl),{"color":"#888888"})
        points.append({"name":row["name"],"country":row["country"],"cluster":int(lbl),"color":li["color"],"x":round(float(coord[0]),4),"y":round(float(coord[1]),4),"price":row["price_usd_kg"],"reliability":row["reliability_score"],"anti_dumping":row["anti_dumping_risk"]})
    return {"mode":"supplier","k":k,"features":features,"silhouette_score":round(sil,4),"inertia":round(float(km.inertia_),2),"elbow":{"k_values":list(range(2,9)),"inertias":inertias},"clusters":cluster_stats,"points":points,"variance_explained":[round(float(v)*100,1) for v in pca.explained_variance_ratio_]}


def run_kmeans_segments(k=4):
    if not ML_AVAILABLE:
        return {"error":"scikit-learn not installed"}
    features = ["annual_demand_mt","growth_rate_pct","price_sensitivity","compliance_priority","speed_priority","volume_per_order_mt"]
    df = pd.DataFrame(CONSTRUCTION_SEGMENTS)
    X = df[features].values
    scaler = StandardScaler()
    X_sc = scaler.fit_transform(X)
    km = KMeans(n_clusters=k, random_state=42, n_init=10)
    labels = km.fit_predict(X_sc)
    sil = float(silhouette_score(X_sc, labels))
    pca = PCA(n_components=2, random_state=42)
    coords = pca.fit_transform(X_sc)
    inertias = []
    for ki in range(2,7):
        km2 = KMeans(n_clusters=ki, random_state=42, n_init=10)
        km2.fit(X_sc)
        inertias.append(round(float(km2.inertia_),2))
    cluster_stats = []
    for c in range(k):
        mask = labels==c
        m = df[mask]
        li = SEGMENT_CLUSTER_LABELS.get(c,{"name":f"Cluster {c}","color":"#888888","description":""})
        cluster_stats.append({"id":int(c),"name":li["name"],"color":li["color"],"description":li["description"],"count":int(mask.sum()),"segments":m["segment"].tolist(),"total_demand_mt":int(m["annual_demand_mt"].sum()),"avg_growth":round(float(m["growth_rate_pct"].mean()),1),"avg_price_sensitivity":round(float(m["price_sensitivity"].mean()),1),"avg_compliance":round(float(m["compliance_priority"].mean()),1)})
    points = []
    for row,lbl,coord in zip(CONSTRUCTION_SEGMENTS,labels,coords):
        li = SEGMENT_CLUSTER_LABELS.get(int(lbl),{"color":"#888888"})
        points.append({"name":row["segment"],"cluster":int(lbl),"color":li["color"],"x":round(float(coord[0]),4),"y":round(float(coord[1]),4),"demand":row["annual_demand_mt"],"growth":row["growth_rate_pct"],"price_sensitivity":row["price_sensitivity"]})
    return {"mode":"segments","k":k,"silhouette_score":round(sil,4),"inertia":round(float(km.inertia_),2),"elbow":{"k_values":list(range(2,7)),"inertias":inertias},"clusters":cluster_stats,"points":points,"variance_explained":[round(float(v)*100,1) for v in pca.explained_variance_ratio_]}


def parse_excel(file_path):
    if not OPENPYXL_AVAILABLE:
        return None
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        result = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data, headers = [], []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if all(v is None for v in row): continue
                clean = [str(v) if v is not None else "" for v in row]
                if i == 0: headers = clean
                else: rows_data.append(clean)
            if headers and rows_data:
                result[sheet_name] = {"headers":headers,"rows":rows_data[:200]}
        return result
    except Exception:
        return None


def parse_csv_file(file_path):
    try:
        with open(file_path,'r',encoding='utf-8-sig') as f:
            rows = list(csv_mod.reader(f))
        if not rows: return None
        return {"Sheet1":{"headers":rows[0],"rows":rows[1:201]}}
    except Exception:
        return None


# ── Django views ──────────────────────────────────────────────────────────────
def index(request):
    ctx = {
        "datasets": DEFAULT_DATASETS,
        "uploaded": UPLOADED_DATASETS,
        "dataset_keys": list(DEFAULT_DATASETS.keys()),
        "ml_available": ML_AVAILABLE,
        "opportunity_data_json": json.dumps(OPPORTUNITY_DATA),
        "india_price_json": json.dumps(INDIA_PRICE_DATA),
        "dimension_weights_json": json.dumps(DIM_WEIGHTS),
        "supplier_count": len(SUPPLIER_DATASET),
        "segment_count": len(CONSTRUCTION_SEGMENTS),
        "kpis": [
            {"label":"Total AU Al Imports 2024","value":"A$1.90B","change":"+4.1% vs 2023","trend":"up","icon":"bi-box-seam","source":"ABS/AAC"},
            {"label":"Extrusions Imported 2024","value":"91,480 t","change":"A$544.9M value","trend":"neutral","icon":"bi-layers","source":"ABS/AAC"},
            {"label":"LME Price Avg 2024","value":"$2,419/t","change":"+7.5% vs 2023","trend":"up","icon":"bi-graph-up-arrow","source":"World Bank/LME"},
            {"label":"China Dumping Duty","value":"ACTIVE","change":"ADN 2023/051 (Sept 2023)","trend":"warning","icon":"bi-exclamation-triangle","source":"Anti-Dumping Commission"},
            {"label":"India Opportunity Score","value":"8.73/10","change":"#1 ranked supplier country","trend":"up","icon":"bi-trophy","source":"NOVO Scoring Model"},
            {"label":"AI Platform Saving","value":"5–8%","change":"A$27M–A$43M on extrusions","trend":"up","icon":"bi-cpu","source":"Project Brief"},
        ],
    }
    return render(request, 'dashboard/index.html', ctx)


@csrf_exempt
@require_http_methods(["POST"])
def upload_file(request):
    if 'file' not in request.FILES:
        return JsonResponse({"error":"No file provided"}, status=400)
    f = request.FILES['file']
    ext = os.path.splitext(f.name)[1].lower()
    if ext not in ['.xlsx','.xls','.csv']:
        return JsonResponse({"error":"Only .xlsx, .xls, .csv supported"}, status=400)
    upload_dir = os.path.join(settings.MEDIA_ROOT,'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, f.name)
    with open(file_path,'wb+') as dest:
        for chunk in f.chunks(): dest.write(chunk)
    parsed = parse_excel(file_path) if ext in ['.xlsx','.xls'] else parse_csv_file(file_path)
    if not parsed:
        return JsonResponse({"error":"Could not parse file."}, status=400)
    key = os.path.splitext(f.name)[0].replace(' ','_').lower()
    UPLOADED_DATASETS[key] = {"label":f.name,"source":"Uploaded","type":"uploaded","sheets":parsed,"filename":f.name}
    return JsonResponse({"success":True,"key":key,"label":f.name,"sheets":list(parsed.keys()),"preview":{s:{"headers":parsed[s]["headers"],"rows":parsed[s]["rows"][:5]} for s in parsed}})


def chart_data(request):
    key = request.GET.get('dataset','au_imports')
    ds = DEFAULT_DATASETS.get(key, DEFAULT_DATASETS['au_imports'])
    return JsonResponse({"label":ds["label"],"source":ds["source"],"type":ds["type"]})


def table_data(request):
    key = request.GET.get('dataset','au_imports')
    if key in UPLOADED_DATASETS:
        ud = UPLOADED_DATASETS[key]
        sheet = request.GET.get('sheet', list(ud['sheets'].keys())[0])
        sd = ud['sheets'].get(sheet,{})
        return JsonResponse({"label":ud["label"],"source":ud["source"],"type":"uploaded","headers":sd.get("headers",[]),"rows":sd.get("rows",[]),"sheets":list(ud['sheets'].keys())})
    ds = DEFAULT_DATASETS.get(key, DEFAULT_DATASETS['au_imports'])
    return JsonResponse({"label":ds["label"],"source":ds["source"],"type":ds["type"],"headers":ds["headers"],"rows":ds["rows"]})


def opportunity_scores(request):
    year = request.GET.get('year','all')
    result = {}
    for country, cdata in OPPORTUNITY_DATA.items():
        years_out = {}
        for y, dims in cdata["years"].items():
            if year != 'all' and str(y) != str(year): continue
            years_out[y] = {**dims,"total":compute_opp_score(dims)}
        result[country] = {"color":cdata["color"],"flag":cdata["flag"],"years":years_out}
    return JsonResponse({"data":result,"weights":DIM_WEIGHTS})


def profit_calc(request):
    india_fob = float(request.GET.get('india_fob',2.30))
    sell_price = float(request.GET.get('sell_price',4.50))
    freight = float(request.GET.get('freight',0.25))
    duty = float(request.GET.get('duty',0.0))
    volume_mt = float(request.GET.get('volume',100))
    aud_usd = float(request.GET.get('aud_usd',0.65))
    landed_usd = india_fob + freight + duty
    landed_aud = landed_usd / aud_usd
    profit_per_kg = sell_price - landed_aud
    margin_pct = (profit_per_kg/sell_price)*100 if sell_price > 0 else 0
    profit_total = profit_per_kg * volume_mt * 1000
    china_landed = (2.95+0.34)/aud_usd
    return JsonResponse({"landed_aud":round(landed_aud,3),"profit_per_kg":round(profit_per_kg,3),"margin_pct":round(margin_pct,1),"profit_total_aud":round(profit_total,0),"advantage_vs_china_aud":round(china_landed-landed_aud,3),"breakeven_sell_price":round(landed_aud,3),"china_landed_aud":round(china_landed,3)})


def run_kmeans(request):
    mode = request.GET.get('mode','supplier')
    k = max(2, min(int(request.GET.get('k',4)), 8))
    result = run_kmeans_supplier(k) if mode == 'supplier' else run_kmeans_segments(k)
    return JsonResponse(result)


@csrf_exempt
def ai_chat(request):
    if request.method != 'POST':
        return JsonResponse({"error":"POST required"}, status=405)
    try:
        body = json.loads(request.body)
        messages = body.get('messages',[])
        context = body.get('context','')
        api_key = body.get('api_key','')
    except Exception:
        return JsonResponse({"error":"Invalid JSON"}, status=400)

    import urllib.request, urllib.error

    system = """You are NOVO, the AI trade intelligence assistant embedded in the NovoAnalytics cross-border aluminium procurement platform for Australian construction firms.

Platform context:
""" + context + """

Key verified facts:
- Australia imports A$1.90B of aluminium annually (2024); A$544.9M is extrusions (ABS/AAC)
- Anti-dumping duties imposed on Chinese aluminium extrusions Sept 2023 (ADN 2023/051)
- India ranked #1 supplier opportunity: score 8.73/10 (NOVO weighted model, 6 dimensions)
- India landed cost is 17-21% cheaper than China after freight (2022-2024 quarterly data)
- K-means clustering (k=4) identified: Premium reliable, Cost-competitive India, High-risk China, Regional niche
- Top Indian suppliers by capacity: Hindalco 2.5M MT, Vedanta 1.9M MT, NALCO 460k MT
- AI-ECTA (Australia-India trade agreement) entered force Dec 2022 — staged duty reductions
- Silhouette score for k=4 supplier clustering: 0.42+ (reasonable cluster separation)

Respond with specific, data-driven answers. Reference cluster names, supplier names, and real figures where relevant. Keep answers concise and actionable."""

    payload = json.dumps({"model":"claude-sonnet-4-6","max_tokens":800,"system":system,"messages":messages}).encode('utf-8')
    req = urllib.request.Request("https://api.anthropic.com/v1/messages", data=payload,
        headers={"Content-Type":"application/json","anthropic-version":"2023-06-01","x-api-key":api_key})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read())
            return JsonResponse({"reply": data['content'][0]['text'] if data.get('content') else "No response"})
    except urllib.error.HTTPError as e:
        return JsonResponse({"error":f"API error {e.code}: {e.read().decode('utf-8','replace')}"}, status=502)
    except Exception as e:
        return JsonResponse({"error":str(e)}, status=502)


@csrf_exempt
def analyse_dataset(request):
    if request.method != 'POST':
        return JsonResponse({"error":"POST required"}, status=405)
    try:
        body = json.loads(request.body)
        dataset_key = body.get('dataset_key','')
        api_key = body.get('api_key','')
    except Exception:
        return JsonResponse({"error":"Invalid JSON"}, status=400)

    ds = DEFAULT_DATASETS.get(dataset_key) or UPLOADED_DATASETS.get(dataset_key)
    if not ds: return JsonResponse({"error":"Dataset not found"}, status=404)

    if 'sheets' in ds:
        sh = list(ds['sheets'].keys())[0]
        headers = ds['sheets'][sh]['headers']
        rows = ds['sheets'][sh]['rows'][:10]
    else:
        headers = ds.get('headers',[])
        rows = ds.get('rows',[])[:10]

    preview = "Columns: " + ", ".join(headers) + "\n"
    for row in rows:
        preview += " | ".join(str(c) for c in row) + "\n"

    import urllib.request, urllib.error
    payload = json.dumps({"model":"claude-sonnet-4-6","max_tokens":1000,"messages":[{"role":"user","content":f"Analyse this dataset for the NOVO aluminium trade platform:\n\nDataset: {ds.get('label',dataset_key)}\nSource: {ds.get('source','')}\n\n{preview}\n\nProvide:\n1. Brief summary (2 sentences)\n2. Three key insights for Australian aluminium procurement from India vs China\n3. One specific recommendation for NovoAnalytics\n4. Data quality notes\n\nBe concise and specific with numbers."}]}).encode('utf-8')
    req = urllib.request.Request("https://api.anthropic.com/v1/messages", data=payload,
        headers={"Content-Type":"application/json","anthropic-version":"2023-06-01","x-api-key":api_key})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read())
            return JsonResponse({"analysis": data['content'][0]['text'] if data.get('content') else "No response","dataset":ds.get('label',dataset_key)})
    except urllib.error.HTTPError as e:
        return JsonResponse({"error":f"API error {e.code}: {e.read().decode('utf-8','replace')}"}, status=502)
    except Exception as e:
        return JsonResponse({"error":str(e)}, status=502)
